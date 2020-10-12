# import sys
import datetime
from flask import current_app
import openpyxl as oxl
from flask_script import Command
from app import db
from app.user_models import User
from app.member.models import UserDetail


class UploadUsers(Command):
    """Upload users from excel or CSV."""
    capture_all_args = True

    def run(self, filename):
        upload_users(filename[0])


def upload_users(filename):
    columns_keys = ['employee_number', 'last_name', 'first_name', 'suffix',
                    'middle_name', 'email']
    wb = oxl.load_workbook(filename)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    some_columns_missing = False
    for i in columns_keys:
        if i not in headers:
            print('\n') if not some_columns_missing else False
            print("Column '{}' not found in upload.".format(i))
            some_columns_missing = True

    if some_columns_missing:
        return

    columns_values = [headers.index(i) + 1 for i in columns_keys]
    columns_dict = dict(zip(columns_keys, columns_values))
    rec_count = 0

    duplicate_rows = []
    employee_numbers = db.session.query(User.employee_number).all()
    employee_number_list = [v[0] for v in employee_numbers]

    rows_to_process = ws.max_row
    for i in range(2, rows_to_process + 1):
        if str(ws.cell(i, columns_dict['employee_number'])
               .value) in employee_number_list:
            duplicate_rows.append(i)
        else:
            user = User(
                email=ws.cell(i, columns_dict['email']).value,
                employee_number=ws.cell(i,
                                        columns_dict['employee_number']).value,
                password=current_app.user_manager.hash_password(
                    current_app.config['DEFAULT_USR_PWD']),
                active=True,
                email_confirmed_at=datetime.datetime.utcnow())
            db.session.add(user)
            userdetail = UserDetail(
                employee_number=ws.cell(i,
                                        columns_dict['employee_number']).value,
                last_name=ws.cell(i, columns_dict['last_name']).value,
                first_name=ws.cell(i, columns_dict['first_name']).value,
                middle_name=ws.cell(i, columns_dict['middle_name']).value,
                suffix=ws.cell(i, columns_dict['suffix']).value)
            db.session.add(userdetail)
            rec_count += 1
            if rec_count > 999:
                db.session.commit()
                rec_count = 0
        db.session.commit()

    if duplicate_rows:
        wb2 = oxl.Workbook()
        ws2 = wb2.active

        ws2.append(headers)
        for i in duplicate_rows:
            values_to_copy = [c.value for c in ws[i]]
            ws2.append(values_to_copy)

        wb2.save('duplicates.xlsx')
        print('\nDuplicate rows saved in \'duplicates.xlsx\'')
    else:
        print('\nDone.')

    # for i in range(2, ws.max_row + 1):
    #     email = ws.cell(i, columns_dict['email'])
    #     employee_number = ws.cell(i, columns_dict['employee_number'])
    #     if not all([email, employee_number]):
    #         continue
    #     employee_exists = db.session.query(User.id)\
    #         .filter_by(employee_number=employee_number).scalar() is not None
    #     if employee_exists:
    #         rows_not_saved.append(i)  # this record at row i is not saved
    #     else:
    #         user = User(
    #             email=ws.cell(i, columns_dict['email']),
    #             employee_number=ws.cell(i, columns_dict['employee_number']),
    #             password=current_app.user_manager.hash_password(
    #                 current_app.config['DEFAULT_USR_PWD']),
    #             active=True,
    #             email_confirmed_at=datetime.datetime.utcnow())
    #         db.session.add(user)
    #         employee_detail_exists = db.session.query(UserDetail.id)\
    #             .filter_by(employee_number=employee_number)\
    #             .scalar() is not None
    #         if employee_detail_exists:
    #             rows_not_saved.append(i)
    #         else:
    #             userdetail = UserDetail(
    #                 employee_number=employee_number,
    #                 last_name=ws.cell(i, columns_dict['last_name']),
    #                 first_name=ws.cell(i, columns_dict['first_name']),
    #                 middle_name=ws.cell(i, columns_dict['middle_name']),
    #                 suffix=ws.cell(i, columns_dict['suffix']))
    #             db.session.add(userdetail)
    #         rec_count += 1
    #         if rec_count > 999:
    #             db.session.commit()
    #             rec_count = 0
    # db.session.commit()
